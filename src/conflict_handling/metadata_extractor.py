"""
Metadata Extractor

Extracts temporal and version metadata from various document types
including emails (PST), PDFs, Office documents, etc.
"""

import os
from datetime import datetime
from typing import Optional
from pathlib import Path

from .models import DocumentMetadata, SourceType


class MetadataExtractor:
    """
    Extract metadata from various document types.

    Supports:
    - Email messages (from PST extraction)
    - PDF documents
    - Office documents (DOCX, XLSX, PPTX)
    - Plain text files
    """

    def extract(
        self,
        file_path: str,
        source_type: Optional[SourceType] = None,
        email_data: Optional[dict] = None
    ) -> DocumentMetadata:
        """
        Extract metadata from a file.

        Args:
            file_path: Path to the file
            source_type: Type of source (auto-detected if None)
            email_data: Pre-extracted email data (for PST extraction)

        Returns:
            DocumentMetadata with all available temporal information
        """
        path = Path(file_path)

        # Auto-detect source type
        if source_type is None:
            source_type = self._detect_source_type(path)

        # Create base metadata
        metadata = DocumentMetadata(
            doc_id=self._generate_doc_id(path),
            source_file=str(path),
            source_type=source_type,
            extraction_date=datetime.now()
        )

        # Extract file system dates
        self._extract_filesystem_dates(path, metadata)

        # Extract type-specific metadata
        if source_type == SourceType.EMAIL and email_data:
            self._extract_email_metadata(email_data, metadata)
        elif source_type == SourceType.PDF:
            self._extract_pdf_metadata(path, metadata)
        elif source_type in [SourceType.DOCX, SourceType.XLSX, SourceType.PPTX]:
            self._extract_office_metadata(path, metadata)

        return metadata

    def extract_from_email(self, email_data: dict) -> DocumentMetadata:
        """
        Extract metadata from email data (from PST extraction).

        Args:
            email_data: Dictionary with email fields:
                - subject, from, to, sent_time, received_time
                - body, attachments, thread_id, etc.

        Returns:
            DocumentMetadata for the email
        """
        metadata = DocumentMetadata(
            doc_id=self._generate_email_id(email_data),
            source_file=email_data.get("source_pst", "unknown"),
            source_type=SourceType.EMAIL,
            extraction_date=datetime.now()
        )

        self._extract_email_metadata(email_data, metadata)

        return metadata

    def _detect_source_type(self, path: Path) -> SourceType:
        """Detect source type from file extension"""
        extension = path.suffix.lower()

        type_map = {
            ".pdf": SourceType.PDF,
            ".docx": SourceType.DOCX,
            ".doc": SourceType.DOCX,
            ".xlsx": SourceType.XLSX,
            ".xls": SourceType.XLSX,
            ".pptx": SourceType.PPTX,
            ".ppt": SourceType.PPTX,
            ".txt": SourceType.TXT,
            ".html": SourceType.HTML,
            ".htm": SourceType.HTML,
            ".eml": SourceType.EMAIL,
            ".msg": SourceType.EMAIL,
        }

        return type_map.get(extension, SourceType.TXT)

    def _generate_doc_id(self, path: Path) -> str:
        """Generate unique document ID"""
        import hashlib
        # Use file path + modification time for uniqueness
        stat = path.stat() if path.exists() else None
        mtime = str(stat.st_mtime) if stat else "0"
        unique_str = f"{path}:{mtime}"
        return hashlib.md5(unique_str.encode()).hexdigest()[:16]

    def _generate_email_id(self, email_data: dict) -> str:
        """Generate unique ID for email"""
        import hashlib
        # Use sender + subject + sent time
        unique_str = (
            f"{email_data.get('from', '')}:"
            f"{email_data.get('subject', '')}:"
            f"{email_data.get('sent_time', '')}"
        )
        return hashlib.md5(unique_str.encode()).hexdigest()[:16]

    def _extract_filesystem_dates(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract dates from file system"""
        if not path.exists():
            return

        stat = path.stat()

        # Creation time (may not be available on all systems)
        try:
            metadata.created_date = datetime.fromtimestamp(stat.st_ctime)
        except (OSError, ValueError):
            pass

        # Modification time
        try:
            metadata.modified_date = datetime.fromtimestamp(stat.st_mtime)
        except (OSError, ValueError):
            pass

    def _extract_email_metadata(
        self,
        email_data: dict,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from email data"""
        # Temporal fields
        if "sent_time" in email_data:
            sent = email_data["sent_time"]
            if isinstance(sent, datetime):
                metadata.sent_date = sent
            elif isinstance(sent, str):
                try:
                    from dateutil import parser
                    metadata.sent_date = parser.parse(sent)
                except (ValueError, TypeError):
                    pass

        if "received_time" in email_data:
            received = email_data["received_time"]
            if isinstance(received, datetime):
                metadata.received_date = received
            elif isinstance(received, str):
                try:
                    from dateutil import parser
                    metadata.received_date = parser.parse(received)
                except (ValueError, TypeError):
                    pass

        # Email-specific fields
        metadata.email_subject = email_data.get("subject")
        metadata.email_from = email_data.get("from")
        metadata.email_to = email_data.get("to", [])
        if isinstance(metadata.email_to, str):
            metadata.email_to = [metadata.email_to]

        metadata.email_thread_id = email_data.get("thread_id") or email_data.get("conversation_id")

        # Author from sender
        metadata.author = email_data.get("from")

        # Check if this is an attachment
        metadata.is_attachment = email_data.get("is_attachment", False)

        # Language detection hint
        if "language" in email_data:
            metadata.language = email_data["language"]

    def _extract_pdf_metadata(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from PDF file"""
        try:
            import pypdf
        except ImportError:
            # pypdf not installed, skip PDF metadata
            return

        try:
            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                info = reader.metadata

                if info is None:
                    return

                # Creation date
                if "/CreationDate" in info:
                    metadata.created_date = self._parse_pdf_date(
                        info["/CreationDate"]
                    )

                # Modification date
                if "/ModDate" in info:
                    metadata.modified_date = self._parse_pdf_date(
                        info["/ModDate"]
                    )

                # Author
                if "/Author" in info:
                    metadata.author = info["/Author"]

                # Title (might contain version info)
                if "/Title" in info:
                    title = info["/Title"]
                    # Try to extract version from title
                    import re
                    version_match = re.search(r"v(?:ersion)?\s*(\d+(?:\.\d+)*)", title, re.I)
                    if version_match:
                        metadata.version = version_match.group(1)

        except Exception:
            # PDF parsing failed, continue without PDF-specific metadata
            pass

    def _parse_pdf_date(self, date_str: str) -> Optional[datetime]:
        """Parse PDF date format (D:YYYYMMDDHHmmSS)"""
        if not date_str:
            return None

        try:
            # Remove 'D:' prefix if present
            if date_str.startswith("D:"):
                date_str = date_str[2:]

            # Parse basic format
            if len(date_str) >= 14:
                return datetime.strptime(date_str[:14], "%Y%m%d%H%M%S")
            elif len(date_str) >= 8:
                return datetime.strptime(date_str[:8], "%Y%m%d")
        except (ValueError, TypeError):
            pass

        return None

    def _extract_office_metadata(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from Office documents (DOCX, XLSX, PPTX)"""
        try:
            from docx import Document
            from openpyxl import load_workbook
        except ImportError:
            # Office libraries not installed
            return

        suffix = path.suffix.lower()

        try:
            if suffix in [".docx", ".doc"]:
                self._extract_docx_metadata(path, metadata)
            elif suffix in [".xlsx", ".xls"]:
                self._extract_xlsx_metadata(path, metadata)
            elif suffix in [".pptx", ".ppt"]:
                self._extract_pptx_metadata(path, metadata)
        except Exception:
            # Office parsing failed
            pass

    def _extract_docx_metadata(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from DOCX file"""
        try:
            from docx import Document

            doc = Document(str(path))
            props = doc.core_properties

            if props.created:
                metadata.created_date = props.created
            if props.modified:
                metadata.modified_date = props.modified
            if props.author:
                metadata.author = props.author
            if props.last_modified_by:
                metadata.last_modified_by = props.last_modified_by
            if props.version:
                metadata.version = props.version
            if props.revision:
                metadata.revision = props.revision

        except Exception:
            pass

    def _extract_xlsx_metadata(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from XLSX file"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True)
            props = wb.properties

            if props.created:
                metadata.created_date = props.created
            if props.modified:
                metadata.modified_date = props.modified
            if props.creator:
                metadata.author = props.creator
            if props.lastModifiedBy:
                metadata.last_modified_by = props.lastModifiedBy
            if props.version:
                metadata.version = props.version

            wb.close()

        except Exception:
            pass

    def _extract_pptx_metadata(
        self,
        path: Path,
        metadata: DocumentMetadata
    ) -> None:
        """Extract metadata from PPTX file"""
        try:
            from pptx import Presentation

            prs = Presentation(str(path))
            props = prs.core_properties

            if props.created:
                metadata.created_date = props.created
            if props.modified:
                metadata.modified_date = props.modified
            if props.author:
                metadata.author = props.author
            if props.last_modified_by:
                metadata.last_modified_by = props.last_modified_by
            if props.version:
                metadata.version = props.version
            if props.revision:
                metadata.revision = props.revision

        except Exception:
            pass


# Convenience function
def extract_metadata(
    file_path: str,
    email_data: Optional[dict] = None
) -> DocumentMetadata:
    """
    Extract metadata from a file.

    Args:
        file_path: Path to the file
        email_data: Pre-extracted email data (for PST)

    Returns:
        DocumentMetadata with temporal information
    """
    extractor = MetadataExtractor()
    return extractor.extract(file_path, email_data=email_data)
